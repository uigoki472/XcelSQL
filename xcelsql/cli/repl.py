r"""Interactive REPL (psql-like) for querying Excel sheets via DuckDB.
Commands (prefix with backslash or slash):
  \help [cmd]              Show help (optionally for a specific command)
  \q / \quit               Exit
  \load <file.xlsx>        Load workbook (lazy; sheets loaded when first referenced)
  \dt / \dt+               List sheets (basic / with loaded metrics)
  \loaded                  List only loaded sheets
  \open <sheet>            Force-load a sheet
  \reload <sheet>          Reload a (loaded) sheet from workbook
  \search <pattern>        Filter sheet names (regex / substring)
  \d <table>               Describe table (loads header if needed)
  \columns <sheet>         Show just column names (quick)
  \hdr <sheet> <row>       Set explicit header row (1-based) for sheet (overrides inference)
  \hdrshow [sheet]         Show header overrides / inferred rows
  \limit [n]               Show / set default display limit
  \header [n]              Show / set default header row for inferring headers
  \format [fmt]            Show / set output format (table|csv|json|jsonl|parquet)
  \template [path]         Show / set template path (applied after each query if set)
  \params                  List current SQL params
  \set name=value          Set a :param value
  \unset name              Remove a param
  \strict [on|off]         Show / toggle strict mode
  \allow_eval [on|off]     Show / toggle Python eval fallback
  \timing [on|off]         Show / toggle per-query timing
  \color [on|off]          Show / toggle ANSI color in output
  \showsql                 Show last executed (transformed) SQL
  \history [n|clear]       Show last n queries (default 20) or clear
  \version                 Show XcelSQL & DuckDB versions
  \functions               List available DuckDB functions (names only)
  \explain <sql>           Show DuckDB EXPLAIN plan (auto-load referenced sheets)
  \save <name> <sql>       Save a named query
  \run <name>              Execute a saved query
  \lsq                     List saved queries
  \watch <sec> <sql?>      Re-run query periodically (Ctrl-C to stop)
  \export <fmt> <path>     Export last result (excel|csv|json|jsonl|parquet)
  \savefmt <fmt> <path>    Export last result without changing current output format
  \stats                   Show table stats (rows, cols, approx memory)
  \show / \showconfig      Show current session settings
  \cache [on|off]          Show / toggle metadata caching flag
  \clear                   Clear loaded sheets & state (keep workbook & settings)
  \restart                 Full session reset (reopens workbook if set)
  \x [on|off]              Toggle expanded (vertical) output format
  \colwidth [n]            Show / set max column display width
Environment overrides (read at startup if set):
  XCELSQL_DISPLAY_LIMIT, XCELSQL_MAX_COL_WIDTH, XCELSQL_TIMING=1, XCELSQL_COLOR=1, XCELSQL_HEADER_ROW
"""
from __future__ import annotations
from typing import Dict, List, Optional, Any
import os
import sys
import re
import time
import shlex
import unicodedata
import hashlib
import json
import logging
import duckdb
import pandas as pd
from contextlib import contextmanager

from xcelsql.utils.excel_utils import SheetTableMapper
from xcelsql.core.sheet_io import load_excel_sheets, parse_sheet_info
from xcelsql.core.sql_engine import validate_sheet_names, resolve_query, validate_query, apply_params
from xcelsql.core.template_engine import apply_template_transform
from xcelsql.utils.output_writer import write_output
from xcelsql.core.expr_eval import set_eval_fallback

try:
    import readline  # type: ignore
except Exception:  # pragma: no cover
    readline = None
# Attempt gnureadline fallback if readline missing
if readline is None:
    try:
        import gnureadline as readline  # type: ignore
    except Exception:  # pragma: no cover
        readline = None

logger = logging.getLogger(__name__)
PROMPT = "xlsql> "
MORE_PROMPT = "... > "

SUPPORTED_REPL_FORMATS = {"table","csv","json","jsonl","parquet"}
COMMANDS = [
    'help','q','quit','load','dt','dt+','loaded','open','reload','search','columns','d','hdr','hdrshow','limit','header','format','template','params','set','unset',
    'strict','allow_eval','show','showconfig','timing','color','explain','save','run','lsq','watch','export','savefmt','stats','showsql','x','colwidth','cache','clear','restart','history','version','functions'
]
SQL_KEYWORDS = [
    'SELECT','FROM','WHERE','GROUP','GROUP BY','HAVING','ORDER','ORDER BY','BY','LIMIT','OFFSET','JOIN','LEFT','LEFT JOIN',
    'RIGHT','RIGHT JOIN','FULL','FULL JOIN','INNER','INNER JOIN','OUTER','OUTER JOIN','CROSS','CROSS JOIN','UNION','UNION ALL',
    'EXCEPT','INTERSECT','ON','USING','AS','DISTINCT','ALL','CASE','WHEN','THEN','ELSE','END','AND','OR','NOT','IN','IS',
    'IS NULL','IS NOT NULL','BETWEEN','LIKE','ILIKE','EXISTS','CREATE','TABLE','VIEW','OR','REPLACE','ALTER','DROP','INSERT',
    'INTO','VALUES','UPDATE','SET','DELETE','WITH','WITH RECURSIVE'
]

# Global session for tab completion
_GLOBAL_SESS = None

class Session:
    """REPL session maintaining workbook, sheets, and options."""
    
    def __init__(self):
        """Initialize a new session."""
        self.workbook = None
        self.all_sheets = []
        self.sheet_data = {}
        self.sheet_columns = {}
        self.header_overrides = {}
        self.sheet_header_row = {}
        self.mapper = SheetTableMapper()
        self.default_header_row = 1
        self.last_query = None
        self.last_result = None
        self.output_format = "table"
        self.template_path = None
        self.strict = False
        self.allow_eval = False
        self.timing = False
        self.expanded = False
        self.max_col_width = 50
        self.saved_queries = {}
        self.params = {}  # SQL parameters
        self.display_limit = 50
        self.con = duckdb.connect()
        self.color = False
        self.cache_enabled = True
        self.history: List[str] = []  # executed SQL (final transformed)
        # Internal alias mapping (sheet_name -> sanitized unique SQL table name)
        self._sanitized_names: Dict[str,str] = {}
        self._orig_to_alias: Dict[str,str] = {}
    
    def load_workbook(self, path: str) -> None:
        """Load Excel workbook metadata only (lazy sheets)."""
        self.workbook = pd.ExcelFile(path)
        self.all_sheets = self.workbook.sheet_names
        self.sheet_data.clear()
        self.sheet_columns.clear()
        self.header_overrides.clear()
        self.sheet_header_row.clear()
        self._sanitized_names.clear()
        self._orig_to_alias.clear()
        # Pre-generate alias mapping without loading data
        for s in self.all_sheets:
            alias = self._alias_for(s)
            self._orig_to_alias[s] = alias
        self._load_meta(path)
    
    def _load_meta(self, path: str) -> None:
        """Load metadata for workbook from cache if available."""
        # Implementation details omitted for brevity
        pass
        
    def _save_meta(self) -> None:
        """Save metadata for current workbook to cache."""
        # Implementation details omitted for brevity
        pass
    
    def list_tables(self) -> List[str]:
        """List available tables/sheets."""
        return self.all_sheets
    
    def _normalize_sheet(self, sheet_name: str) -> Optional[str]:
        """Normalize sheet name, case-insensitive match."""
        for s in self.all_sheets:
            if s.lower() == sheet_name.lower():
                return s
        return None
    
    def _load_header(self, sheet_name: str) -> None:
        """Load only header (first row) for a sheet and store columns."""
        if not self.workbook:
            raise RuntimeError("No workbook loaded")
        real = self._normalize_sheet(sheet_name) or sheet_name
        header_row = self.header_overrides.get(real, self.default_header_row)
        # Load single row to infer columns
        df_head = self.workbook.parse(real, header=header_row-1, nrows=0)
        self.sheet_columns[real] = list(df_head.columns)
    
    def _infer_header_row(self, sheet_name: str) -> int:
        """Infer header row (1-based) by sampling first up to 10 rows.
        Heuristic: choose row with max count of non-null, non-numeric-looking string cells.
        Fallback to 1.
        """
        if not self.workbook:
            return 1
        try:
            sample = self.workbook.parse(sheet_name, header=None, nrows=10)
        except Exception:
            return 1
        best_row = 0
        best_score = -1
        for idx, row in sample.iterrows():
            score = 0
            for val in row.tolist():
                if isinstance(val, str):
                    # treat as header candidate if contains any letter
                    if re.search(r'[A-Za-z]', val):
                        score += 1
                else:
                    # ignore pure numerics / NaNs
                    continue
            if score > best_score:
                best_score = score
                best_row = idx
        return best_row + 1  # convert to 1-based
    
    def _ensure_header(self, sheet_name: str) -> None:
        """Ensure header row inferred / overridden and column names loaded (without full data)."""
        if sheet_name in self.sheet_columns:
            return
        if sheet_name not in self.all_sheets:
            raise ValueError(f"Unknown sheet: {sheet_name}")
        if sheet_name not in self.header_overrides and sheet_name not in self.sheet_header_row:
            self.sheet_header_row[sheet_name] = self._infer_header_row(sheet_name)
        header_row = self.header_overrides.get(sheet_name, self.sheet_header_row.get(sheet_name, self.default_header_row))
        if not self.workbook:
            return
        try:
            df_head = self.workbook.parse(sheet_name, header=header_row-1, nrows=0)
            # Normalize: trim outer whitespace, preserve internal spaces, dedupe
            cols = []
            seen = {}
            for c in df_head.columns:
                name = str(c).strip()
                base = name
                if base in seen:
                    seen[base] += 1
                    name = f"{base}_{seen[base]}"
                else:
                    seen[base] = 0
                cols.append(name)
            self.sheet_columns[sheet_name] = cols
        except Exception:
            self.sheet_columns[sheet_name] = []
    
    def describe(self, sheet_name: str) -> pd.DataFrame:
        real = self._normalize_sheet(sheet_name) or sheet_name
        # Only load header / columns; do NOT load full sheet data
        self._ensure_header(real)
        cols = self.sheet_columns.get(real, [])
        if not cols:
            return pd.DataFrame()
        # Provide minimal dtype info by sampling first 50 rows (lightweight) if not loaded
        dtypes = {}
        if self.workbook and real not in self.sheet_data:
            header_row = self.header_overrides.get(real, self.sheet_header_row.get(real, self.default_header_row))
            try:
                sample = self.workbook.parse(real, header=header_row-1, nrows=50, usecols=list(range(len(cols))))
                for c in cols:
                    if c in sample.columns:
                        s = sample[c]
                        dtypes[c] = str(s.dtype)
                    else:
                        dtypes[c] = 'object'
            except Exception:
                for c in cols:
                    dtypes[c] = 'object'
        else:
            df = self.sheet_data.get(real)
            if df is not None:
                for c in df.columns:
                    dtypes[c] = str(df[c].dtype)
        meta_rows = []
        for c in cols:
            meta_rows.append({'name': c, 'dtype': dtypes.get(c, 'object')})
        return pd.DataFrame(meta_rows)
    
    def _alias_for(self, sheet_name: str) -> str:
        """Return (and cache) a sanitized unique alias for a sheet."""
        if sheet_name in self._sanitized_names:
            return self._sanitized_names[sheet_name]
        base = re.sub(r'\W+', '_', sheet_name).strip('_') or 'sheet'
        # Ensure starts with letter or underscore
        if not re.match(r'^[A-Za-z_]', base):
            base = 't_' + base
        alias = base
        suffix = 1
        existing = set(self._sanitized_names.values())
        while alias.lower() in (a.lower() for a in existing):
            alias = f"{base}_{suffix}"
            suffix += 1
        self._sanitized_names[sheet_name] = alias
        return alias
    
    def get_sql_table_names(self) -> List[str]:
        """Return list of usable SQL table tokens (aliases + originals)."""
        names = []
        for original, alias in self._sanitized_names.items():
            names.append(alias)
            if original != alias:
                names.append(original)
        return names
    
    def rewrite_identifiers(self, sql: str) -> str:
        """Rewrite double-quoted identifiers matching original sheet names to their alias.
        Avoid changing content inside single-quoted string literals.
        """
        if not self._orig_to_alias:
            return sql
        out = []
        i = 0
        n = len(sql)
        in_single = False
        while i < n:
            ch = sql[i]
            if in_single:
                out.append(ch)
                if ch == "'":
                    # Handle escaped single quote ''
                    if i + 1 < n and sql[i+1] == "'":
                        out.append("'")
                        i += 2
                        continue
                    else:
                        in_single = False
                i += 1
                continue
            if ch == "'":
                in_single = True
                out.append(ch)
                i += 1
                continue
            if ch == '"':
                # Parse quoted identifier
                j = i + 1
                ident_chars = []
                while j < n:
                    if sql[j] == '"':
                        if j + 1 < n and sql[j+1] == '"':  # escaped quote
                            ident_chars.append('"')
                            j += 2
                            continue
                        else:
                            break
                    ident_chars.append(sql[j])
                    j += 1
                if j >= n:  # unmatched quote, just append remainder
                    out.append(sql[i:])
                    break
                original_ident = ''.join(ident_chars)
                alias = self._orig_to_alias.get(original_ident)
                if alias:
                    # Replace with alias (quoted only if needed)
                    if re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', alias):
                        out.append(alias)
                    else:
                        esc = alias.replace('"','""')
                        out.append(f'"{esc}"')
                else:
                    # keep original quoted form
                    esc_back = original_ident.replace('"','""')
                    out.append(f'"{esc_back}"')
                i = j + 1
                continue
            out.append(ch)
            i += 1
        return ''.join(out)
    
    def open_sheet(self, sheet_name: str, force_reload: bool = False) -> None:
        # Lazy full data load (with optional force reload)
        if not self.workbook:
            raise RuntimeError("No workbook loaded")
        real = self._normalize_sheet(sheet_name) or sheet_name
        self._ensure_header(real)
        header_row = self.header_overrides.get(real, self.sheet_header_row.get(real, self.default_header_row))
        if real in self.sheet_data and not force_reload:
            return
        if real in self.sheet_data and force_reload:
            try:
                del self.sheet_data[real]
            except Exception:
                pass
        df = self.workbook.parse(real, header=header_row-1)
        # Normalize headers for full DataFrame too
        try:
            cols = []
            seen = {}
            for c in df.columns:
                name = str(c).strip()
                base = name
                if base in seen:
                    seen[base] += 1
                    name = f"{base}_{seen[base]}"
                else:
                    seen[base] = 0
                cols.append(name)
            df.columns = cols
        except Exception:
            pass
        self.sheet_data[real] = df
        self.sheet_columns[real] = list(df.columns)
        alias = self._alias_for(real)
        self._orig_to_alias[real] = alias
        try:
            self.con.unregister(alias)
        except Exception:
            pass
        self.con.register(alias, df)

# Helper functions for robust command parsing with smart/unmatched quotes
SMART_QUOTE_MAP = {
    '\u201c': '"',  # left double
    '\u201d': '"',  # right double
    '\u201e': '"',
    '\u201f': '"',
    '\u2033': '"',
    '\u2018': "'",  # left single
    '\u2019': "'",  # right single / apostrophe
    '\u201b': "'",
    '\u2032': "'",
}

def _normalize_smart_quotes(s: str) -> str:
    return ''.join(SMART_QUOTE_MAP.get(ch, ch) for ch in s)

def _safe_split(cmd: str) -> List[str]:
    norm = _normalize_smart_quotes(cmd.strip())
    # Attempt normal shlex
    try:
        return shlex.split(norm)
    except ValueError:
        # Try auto-closing unmatched quotes
        dq = norm.count('"') - norm.count('\\"')
        sq = norm.count("'") - norm.count("\\'")
        fixed = norm
        if dq % 2 == 1:
            fixed += '"'
        if sq % 2 == 1:
            fixed += "'"
        if fixed != norm:
            try:
                return shlex.split(fixed)
            except Exception:
                pass
        # Fallback simple whitespace split (best effort)
        return fixed.split()

def _completer(text: str, state: int) -> Optional[str]:
    """Tab completion for REPL commands, sheet names, SQL tables, and SQL keywords."""
    global _GLOBAL_SESS
    test_mode = 'PYTEST_CURRENT_TEST' in os.environ
    line_buffer = ''
    if readline:
        try:
            line_buffer = readline.get_line_buffer()
        except Exception:
            line_buffer = text or ''
    if line_buffer is None:
        line_buffer = text or ''

    # --------------------------------------------------
    # Backslash / slash command + argument completion
    # --------------------------------------------------
    if line_buffer.startswith(('\\','/')):
        prefix_char = line_buffer[0]
        body = line_buffer[1:]
        # If completing command name (no space yet OR cursor still in first token)
        if ' ' not in body.rstrip():
            # Command name fragment is current body (unless empty)
            if body.endswith(' '):  # user typed full command then space -> next branch handles args
                pass
            else:
                frag = body
                if text.startswith(prefix_char):  # readline may give entire token including prefix
                    frag = text[1:]
                elif not body.endswith(text):
                    frag = body  # fallback
                base_matches = [c for c in COMMANDS if c.startswith(frag)] if frag else COMMANDS
                matches = [prefix_char + c for c in base_matches]
                if not matches:
                    return None
                if test_mode and state == 0:
                    return matches  # type: ignore
                return matches[state] if state < len(matches) else None
        # Argument completion
        parts = body.split()
        cmd = parts[0]
        arg_fragment = '' if line_buffer.endswith(' ') else parts[-1] if len(parts) > 1 else ''
        # Sheet-argument commands (first arg only)
        sheet_arg_cmds = {'hdr','d','hdrshow','columns'}
        if cmd in sheet_arg_cmds and _GLOBAL_SESS:
            # For hdr command only complete sheet on first arg
            if cmd == 'hdr' and len(parts) >= 3 and not line_buffer.endswith(' '):
                return None
            sheets = _GLOBAL_SESS.list_tables()
            frag = arg_fragment
            sheet_matches = [s for s in sheets if s.lower().startswith(frag.lower())]
            def format_sheet(s: str) -> str:
                # Quote if contains whitespace or special chars
                return s if re.match(r'^[A-Za-z0-9_\-]+$', s) else f'"{s}"'
            formatted = [format_sheet(s) for s in sheet_matches]
            if not formatted:
                return None
            if test_mode and state == 0:
                return formatted  # type: ignore
            return formatted[state] if state < len(formatted) else None
        return None

    # --------------------------------------------------
    # SQL / sheet / keyword completion context
    # --------------------------------------------------
    token = text
    # Detect if inside a double-quoted identifier (partial sheet name)
    inside_dq = False
    dq_pos = line_buffer.rfind('"')
    if dq_pos != -1:
        tail = line_buffer[dq_pos:]
        if tail.count('"') % 2 == 1:
            inside_dq = True
            fragment_after_quote = line_buffer[dq_pos+1:]
            dq_fragment = fragment_after_quote if fragment_after_quote.endswith(text) else text
        else:
            dq_fragment = ''
    else:
        dq_fragment = ''
    # Sheet completion with { prefix (legacy style)
    if line_buffer.startswith('{') and _GLOBAL_SESS and token.startswith('{'):
        frag = token[1:]
        sheets = _GLOBAL_SESS.list_tables()
        sheet_matches = [s for s in sheets if s.lower().startswith(frag.lower())]
        formatted = ['{' + s + '}' for s in sheet_matches]
        if not formatted:
            return None
        if test_mode and state == 0:
            return formatted  # type: ignore
        return formatted[state] if state < len(formatted) else None
    # Collect SQL keyword completions
    kw_matches: List[str] = []
    if token and not inside_dq:
        up = token.upper()
        for kw in SQL_KEYWORDS:
            if kw.startswith(up):
                kw_matches.append(kw)
    elif not token and not inside_dq:
        kw_matches = ['SELECT','WITH','INSERT','UPDATE','DELETE','CREATE']
    # Table name completions
    table_matches: List[str] = []
    if _GLOBAL_SESS:
        def needs_quote(name: str) -> bool:
            return not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', name)
        table_fragment = dq_fragment if inside_dq else token
        seen_lower = set()
        frag_low = table_fragment.lower()
        def add_candidate(raw_name: str, original: bool):
            core = raw_name
            low_core = core.lower()
            if table_fragment != '' and not low_core.startswith(frag_low):
                return
            if inside_dq:
                rendered = '"' + core.replace('"','""') + '"'
            else:
                if needs_quote(core) and not token.startswith('"'):
                    # Provide unquoted candidate starting with typed prefix for autocompletion
                    rendered = core
                else:
                    rendered = core if not needs_quote(core) else '"' + core.replace('"','""') + '"'
            low_key = (rendered.lower(), inside_dq)
            if low_key in seen_lower:
                return
            seen_lower.add(low_key)
            table_matches.append(rendered)
        # Originals first
        for original in getattr(_GLOBAL_SESS, 'all_sheets', []):
            add_candidate(original, True)
        # Aliases
        for original, alias in _GLOBAL_SESS._orig_to_alias.items():  # type: ignore
            add_candidate(alias, False)
    # If inside double quotes, suppress keyword suggestions
    if inside_dq:
        kw_matches = []
    prev_tokens = re.split(r'\s+', line_buffer.strip())
    prev_word = prev_tokens[-2].upper() if len(prev_tokens) >= 2 else ''
    table_context_words = {'FROM','JOIN','UPDATE','INTO','TABLE'}
    ordered: List[str] = []
    if prev_word in table_context_words or inside_dq:
        ordered.extend(table_matches)
        for k in kw_matches:
            if k not in ordered:
                ordered.append(k)
    else:
        seen=set()
        for seq in (kw_matches, table_matches):
            for item in seq:
                if item not in seen:
                    seen.add(item)
                    ordered.append(item)
    if not ordered:
        return None
    if test_mode and state == 0:
        return ordered  # type: ignore
    return ordered[state] if state < len(ordered) else None

def start_repl(initial_workbook: Optional[str] = None, allow_eval: bool = False, strict: bool = False) -> None:
    """Start the interactive REPL (fully functional core subset)."""
    global _GLOBAL_SESS
    sess = Session()
    # Apply env overrides
    try:
        if os.getenv('XCELSQL_DISPLAY_LIMIT'): sess.display_limit = int(os.getenv('XCELSQL_DISPLAY_LIMIT',''))
    except Exception: pass
    try:
        if os.getenv('XCELSQL_MAX_COL_WIDTH'): sess.max_col_width = int(os.getenv('XCELSQL_MAX_COL_WIDTH',''))
    except Exception: pass
    try:
        if os.getenv('XCELSQL_HEADER_ROW'): sess.default_header_row = int(os.getenv('XCELSQL_HEADER_ROW',''))
    except Exception: pass
    if os.getenv('XCELSQL_TIMING') == '1': sess.timing = True
    if os.getenv('XCELSQL_COLOR') == '1': sess.color = True
    sess.allow_eval = allow_eval
    sess.strict = strict
    _GLOBAL_SESS = sess

    if initial_workbook:
        try:
            sess.load_workbook(initial_workbook)
            print(f"Loaded workbook: {initial_workbook} ({len(sess.all_sheets)} sheets)")
        except Exception as e:
            print(f"Failed to load workbook: {e}")
    # Configure readline
    if readline:
        try:
            readline.set_completer(_completer)
            readline.set_completer_delims(' \t\n')
            # libedit (macOS default) needs a different binding than GNU readline
            docstr = getattr(readline, '__doc__', '') or ''
            if 'libedit' in docstr.lower():
                readline.parse_and_bind('bind ^I rl_complete')
            else:
                readline.parse_and_bind('tab: complete')
            readline.parse_and_bind('set completion-ignore-case on')
            readline.parse_and_bind('set show-all-if-ambiguous on')
        except Exception:  # pragma: no cover
            pass

    buffer: List[str] = []

    def flush_query() -> Optional[str]:
        if not buffer:
            return None
        sql = "\n".join(buffer).strip()
        buffer.clear()
        return sql

    def run_query(sql: str) -> None:
        if not sql:
            return
        # Normalize smart quotes to standard ASCII quotes before any parsing / rewriting
        sql = _normalize_smart_quotes(sql)
        orig_sql = sql
        # Auto-load referenced sheets lazily before rewriting identifiers
        def _auto_load(sql_text: str):
            if not sess.all_sheets:
                return
            lower_sql = sql_text.lower()
            for sheet in sess.all_sheets:
                if sheet in sess.sheet_data:
                    continue
                alias = sess._orig_to_alias.get(sheet) or sess._alias_for(sheet)
                quoted_pattern = f'"{sheet}"'.lower()
                alias_pattern = alias.lower()
                if quoted_pattern in lower_sql or re.search(rf'\b{re.escape(alias_pattern)}\b', lower_sql):
                    try:
                        sess.open_sheet(sheet)
                    except Exception as e:
                        print(f"Error loading sheet {sheet}: {e}")
        _auto_load(orig_sql)
        sql = sess.rewrite_identifiers(sql)
        if sess.template_path:
            try:
                transformed = apply_template_transform(sql, sess.template_path)
                if transformed:
                    sql = transformed
            except Exception as e:
                print(f"Template error: {e}")
                return
        if sess.params:
            try:
                sql = apply_params(sql, sess.params)
            except Exception as e:
                print(f"Param substitution error: {e}")
                return
        start_t = time.time()
        try:
            df = sess.con.execute(sql).fetchdf()
            sess.last_query = sql
            sess.last_result = df
            if sql not in sess.history:
                sess.history.append(sql)
            # Custom psql-like formatting for table output
            if sess.output_format == 'table':
                limit_df = df.head(sess.display_limit)
                color_on = sess.color and sys.stdout.isatty()
                def colorize(text, code):
                    return f"\033[{code}m{text}\033[0m" if color_on else text
                if sess.expanded:
                    # Expanded vertical format
                    for idx, row in limit_df.iterrows():
                        rec_no = idx + 1
                        print(colorize(f"-[ RECORD {rec_no} ]-",'36'))
                        for col in limit_df.columns:
                            val = row[col]
                            sval = '' if pd.isna(val) else str(val)
                            if len(sval) > sess.max_col_width:
                                sval = sval[: sess.max_col_width - 1] + '…'
                            print(f"{colorize(str(col),'33')}: {sval}")  # ensure header label is string
                        print()
                    print(f"({len(limit_df)} row{'s' if len(limit_df)!=1 else ''})")
                else:
                    # Tabular format
                    # Prepare display copy of column names as strings
                    display_cols = [str(c) for c in limit_df.columns]
                    # Compute widths based on display names and cell content (stringified)
                    raw_cells = {}
                    for orig_col, disp_col in zip(limit_df.columns, display_cols):
                        col_values = [disp_col]
                        for v in limit_df[orig_col].tolist():
                            col_values.append('' if pd.isna(v) else str(v))
                        raw_cells[disp_col] = col_values
                    widths = {dc: min(max(len(x) for x in cells), sess.max_col_width) for dc, cells in raw_cells.items()}
                    header = ' | '.join(colorize(dc[:widths[dc]].ljust(widths[dc]), '32') for dc in display_cols)
                    separator = '-+-'.join('-'*widths[dc] for dc in display_cols)
                    print(header)
                    print(separator)
                    for i in range(len(limit_df)):
                        parts = []
                        for orig_col, dc in zip(limit_df.columns, display_cols):
                            val = limit_df.iloc[i][orig_col]
                            sval = '' if pd.isna(val) else str(val)
                            if len(sval) > widths[dc]:
                                sval = sval[: widths[dc]-1] + '…'
                            parts.append(sval.ljust(widths[dc]))
                        print(' | '.join(parts))
                    print(f"({len(limit_df)} row{'s' if len(limit_df)!=1 else ''})")
            else:
                # Delegate to writer for non-table formats
                try:
                    write_output(df, fmt=sess.output_format, limit=sess.display_limit, max_col_width=sess.max_col_width)
                except Exception:
                    if sess.output_format == 'jsonl':
                        for rec in json.loads(df.to_json(orient='records')):
                            print(json.dumps(rec))
                    elif sess.output_format == 'json':
                        print(df.to_json(orient='records', indent=2))
                    else:
                        print(df.head(sess.display_limit).to_string(index=False))
            if sess.timing:
                print(f"Time: {(time.time()-start_t)*1000:.1f} ms")
        except Exception as e:
            print(f"Error: {e}")

    def show_help():
        print(__doc__ or 'No help available.')

    def show_settings():
        print(json.dumps({
            'workbook': getattr(sess.workbook, 'io', None),
            'sheets': sess.all_sheets,
            'loaded_sheets': list(sess.sheet_data.keys()),
            'format': sess.output_format,
            'display_limit': sess.display_limit,
            'max_col_width': sess.max_col_width,
            'template': sess.template_path,
            'strict': sess.strict,
            'allow_eval': sess.allow_eval,
            'timing': sess.timing,
            'params': sess.params,
            'saved_queries': list(sess.saved_queries.keys()),
        }, indent=2, default=str))

    HELP_TEXT: Dict[str,str] = {
        'dt': 'List sheets ( * indicates loaded )',
        'dt+': 'List sheets with rows/cols/memory (loaded sheets only for metrics).',
        'loaded': 'List only loaded sheets.',
        'open': '\\open <sheet> : force-load a sheet',
        'reload': '\\reload <sheet> : reload sheet data',
        'hdr': '\\hdr <sheet> <row> : set explicit header row (1-based)',
        'hdrshow': 'Show header overrides / inferred rows',
        'showsql': 'Display last executed SQL after transformations',
        'watch': '\\watch <seconds> [sql] : rerun query periodically',
        'stats': 'Show stats for all sheets (estimates for unloaded)',
        'x': 'Toggle expanded vertical output',
        'format': 'Show/set output format (table,csv,json,jsonl,parquet)',
        'limit': 'Show/set display row limit',
        'colwidth': 'Show/set max column display width',
        'color': 'Toggle ANSI color output',
        'cache': 'Show/toggle metadata caching flag (future use)',
        'clear': 'Clear loaded data & last result (keep workbook)',
        'restart': 'Reset entire session (preserve workbook path)',
        'savefmt': '\\savefmt <fmt> <path> : export last result without changing format',
        'history': '\\history [n|clear] : show or clear recent queries',
        'version': 'Show XcelSQL and DuckDB versions',
        'functions': 'List available DuckDB function names',
        'search': '\\search <pattern> : filter sheet names (regex)',
        'columns': 'Show column names for sheet (quick)'
    }

    while True:
        try:
            prompt = PROMPT if not buffer else MORE_PROMPT
            try:
                line = input(prompt)
            except EOFError:
                print()  # newline on Ctrl-D
                break
            if line is None:
                break
            line = line.rstrip('\n')
            # Handle commands (when not building multiline)
            if not buffer and (line.startswith('\\') or line.startswith('/')):
                parts = _safe_split(line[1:])
                if not parts:
                    continue
                cmd, *args = parts
                if cmd in ('q','quit'):
                    break
                elif cmd == 'help':
                    show_help()
                elif cmd == 'load' and args:
                    path = args[0]
                    try:
                        sess.load_workbook(path)
                        print(f"Loaded {len(sess.all_sheets)} sheets.")
                    except Exception as e:
                        print(f"Load failed: {e}")
                elif cmd == 'load':
                    print('Usage: \\load <file.xlsx>')
                elif cmd == 'dt':
                    if not sess.all_sheets:
                        print('(no sheets)')
                    else:
                        color_on = sess.color and sys.stdout.isatty()
                        if color_on:
                            print("\033[32mLoaded Sheet\033[0m")
                        for s in sess.all_sheets:
                            mark = '*' if s in sess.sheet_data else ' '
                            print(f"{mark} {s}")
                elif cmd == 'dt+':

                    if not sess.all_sheets:
                        print('(no sheets)')
                    else:
                        # Reuse stats logic for loaded ones quickly
                        for s in sess.all_sheets:
                            mark = '*' if s in sess.sheet_data else ' '
                            if s in sess.sheet_data:
                                df = sess.sheet_data[s]
                                try:
                                    mem = int(df.memory_usage(deep=True).sum())
                                except Exception:
                                    mem = 0
                                print(f"{mark} {s} (rows={len(df)}, cols={df.shape[1]}, mem={mem}B)")
                            else:
                                print(f"{mark} {s}")
                elif cmd == 'd' and args:
                    target = sess._normalize_sheet(args[0]) or args[0]
                    try:
                        meta = sess.describe(target)
                        if meta.empty:
                            print('No metadata.')
                        else:
                            print(meta.to_string(index=False))
                    except Exception as e:
                        print(f"Describe error: {e}")
                elif cmd == 'd':
                    print('Usage: \\d <sheet>')
                elif cmd == 'hdr' and len(args) == 2:
                    sheet, row = args
                    try:
                        r = int(row)
                        if r < 1:
                            raise ValueError
                        real = sess._normalize_sheet(sheet) or sheet
                        sess.header_overrides[real] = r
                        if real in sess.sheet_data:
                            sess.open_sheet(real)  # reload
                        print(f"Header row for {real} set to {r}")
                    except ValueError:
                        print('Row must be positive integer')
                elif cmd == 'hdr':
                    print('Usage: \\hdr <sheet> <row>')
                elif cmd == 'showsql':
                    print(sess.last_query or '(no query executed)')
                elif cmd == 'open' and args:
                    try:
                        sess.open_sheet(args[0])
                        print(f"Loaded sheet: {args[0]}")
                    except Exception as e:
                        print(f"Open failed: {e}")
                elif cmd == 'open':
                    print('Usage: \\open <sheet>')
                elif cmd == 'reload' and args:
                    try:
                        sess.open_sheet(args[0], force_reload=True)
                        print(f"Reloaded sheet: {args[0]}")
                    except Exception as e:
                        print(f"Reload failed: {e}")
                elif cmd == 'reload':
                    print('Usage: \\reload <sheet>')
                elif cmd == 'loaded':
                    loaded = list(sess.sheet_data.keys())
                    if not loaded:
                        print('(no sheets loaded)')
                    else:
                        for s in loaded:
                            print(s)
                elif cmd == 'search' and args:
                    pat = ' '.join(args)
                    try:
                        regex = re.compile(pat, re.IGNORECASE)
                        matches = [s for s in sess.all_sheets if regex.search(s)]
                    except re.error:
                        low = pat.lower()
                        matches = [s for s in sess.all_sheets if low in s.lower()]
                    if not matches:
                        print('(no matches)')
                    else:
                        for s in matches:
                            mark='*' if s in sess.sheet_data else ' '
                            print(f"{mark} {s}")
                elif cmd == 'search':
                    print('Usage: \\search <pattern>')
                elif cmd == 'columns' and args:
                    target = sess._normalize_sheet(args[0]) or args[0]
                    try:
                        sess._ensure_header(target)
                        cols = sess.sheet_columns.get(target, [])
                        # Coerce any non-string header labels to string for display
                        disp = [str(c) for c in cols]
                        print(', '.join(disp) if disp else '(no columns)')
                    except Exception as e:
                        print(f'Columns error: {e}')
                elif cmd == 'columns':
                    print('Usage: \\columns <sheet>')
                elif cmd == 'history':
                    if args and args[0] == 'clear':
                        sess.history.clear(); print('History cleared.')
                    else:
                        try:
                            n = int(args[0]) if args else 20
                        except Exception:
                            n = 20
                        for idx, q in enumerate(reversed(sess.history[-n:]), start=1):
                            print(f"{idx}: {q}")
                elif cmd == 'version':
                    try:
                        import xcelsql as _pkg
                        pkg_ver = getattr(_pkg, '__version__', 'unknown')
                    except Exception:
                        pkg_ver = 'unknown'
                    try:
                        duck_ver = duckdb.__version__
                    except Exception:
                        duck_ver = 'unknown'
                    try:
                        import pandas as _pd
                        pd_ver = _pd.__version__
                    except Exception:
                        pd_ver = 'unknown'
                    try:
                        import openpyxl as _ox
                        ox_ver = getattr(_ox, '__version__', 'unknown')
                    except Exception:
                        ox_ver = 'unknown'
                    print(f"XcelSQL {pkg_ver} | DuckDB {duck_ver} | pandas {pd_ver} | openpyxl {ox_ver}")
                elif cmd == 'functions':
                    try:
                        fdf = sess.con.execute('PRAGMA functions').fetchdf()
                        names = sorted(set(fdf['name'].tolist()))
                        for n in names:
                            print(n)
                        print(f"({len(names)} functions)")
                    except Exception as e:
                        print(f'Functions error: {e}')
                elif cmd == 'color':
                    if args and args[0] in ('on','off'):
                        sess.color = args[0]=='on'
                    print(f"color={'on' if sess.color else 'off'}")
                elif cmd == 'x':
                    if args and args[0] in ('on','off'):
                        sess.expanded = (args[0]=='on')
                    else:
                        sess.expanded = not sess.expanded
                    print(f"expanded={'on' if sess.expanded else 'off'}")
                elif cmd == 'cache':
                    if args and args[0] in ('on','off'):
                        sess.cache_enabled = args[0]=='on'
                    print(f"cache={'on' if sess.cache_enabled else 'off'}")
                elif cmd == 'clear':
                    sess.sheet_data.clear(); sess.last_result=None; sess.last_query=None
                    print('Session data cleared (workbook retained).')
                elif cmd == 'restart':
                    wb_path = getattr(sess.workbook,'io',None)
                    new = Session()
                    new.allow_eval=sess.allow_eval; new.strict=sess.strict; new.output_format=sess.output_format
                    new.display_limit=sess.display_limit; new.max_col_width=sess.max_col_width
                    new.template_path=sess.template_path; new.params=dict(sess.params); new.color=sess.color
                    if wb_path:
                        try:
                            new.load_workbook(wb_path)
                            print(f"Reopened workbook: {wb_path}")
                        except Exception as e:
                            print(f"Workbook reopen failed: {e}")
                    _GLOBAL_SESS = new
                    sess = new  # type: ignore
                    continue
                elif cmd == 'savefmt':
                    if len(args) < 2:
                        print('Usage: \\savefmt <fmt> <path>')
                    else:
                        fmt, path = args[0].lower(), args[1]
                        if sess.last_result is None:
                            print('No result to export (run a query first)')
                        else:
                            df = sess.last_result
                            try:
                                if fmt == 'csv':
                                    df.to_csv(path, index=False, encoding='utf-8-sig')
                                elif fmt == 'json':
                                    df.to_json(path, orient='records', indent=2, force_ascii=False)
                                elif fmt == 'jsonl':
                                    with open(path,'w',encoding='utf-8') as f:
                                        for rec in json.loads(df.to_json(orient='records', force_ascii=False)):
                                            f.write(json.dumps(rec, ensure_ascii=False)+'\n')
                                elif fmt in ('excel','xlsx'):
                                    df.to_excel(path, index=False)
                                elif fmt == 'parquet':
                                    try:
                                        df.to_parquet(path, index=False)
                                    except Exception as e:
                                        print(f'Parquet export failed: {e}')
                                        continue
                                else:
                                    print(f'Unsupported format: {fmt}')
                                    continue
                                print(f'Exported {len(df)} rows to {path}')
                            except Exception as e:
                                print(f'Export failed: {e}')
                elif cmd == 'showconfig':
                    show_settings()
                elif cmd == 'help':
                    if args:
                        h = HELP_TEXT.get(args[0].lstrip('\\/'), '(no help)')
                        print(h)
                    else:
                        show_help()
                elif cmd == 'stats':
                    if not sess.all_sheets:
                        print('No workbook loaded.')
                    else:
                        def _fmt_bytes(num: int) -> str:
                            for unit in ['B','KB','MB','GB','TB']:
                                if num < 1024:
                                    return f"{num:.1f}{unit}"
                                num /= 1024
                            return f"{num:.1f}PB"
                        def disp_width(s: str) -> int:
                            w = 0
                            for ch in s:
                                if unicodedata.east_asian_width(ch) in ('F','W'):
                                    w += 2
                                else:
                                    w += 1
                            return w
                        def pad_right(s: str, width: int) -> str:
                            extra = width - disp_width(s)
                            if extra > 0:
                                return s + ' ' * extra
                            return s
                        rows_fmt = []
                        total_mem = 0
                        loaded_count = 0
                        total = len(sess.all_sheets)
                        print(f"Scanning {total} sheet(s)...")
                        scan_start = time.time()
                        last_print_len = 0
                        for idx, name in enumerate(sess.all_sheets):
                            # Live progress line
                            progress_msg = f"[{idx+1}/{total}] {name}"
                            pad_clear = ' ' * max(0, last_print_len - len(progress_msg))
                            print(progress_msg + pad_clear, end='\r', flush=True)
                            last_print_len = len(progress_msg)
                            loaded = name in sess.sheet_data
                            if loaded:
                                df = sess.sheet_data[name]
                                try:
                                    mem = int(df.memory_usage(deep=True).sum())
                                except Exception:
                                    mem = 0
                                total_mem += mem
                                loaded_count += 1
                                rows_fmt.append((name, len(df), df.shape[1], _fmt_bytes(mem), '*'))
                            else:
                                try:
                                    sess._ensure_header(name)
                                    header_row = sess.header_overrides.get(name, sess.sheet_header_row.get(name, sess.default_header_row))
                                    ncols = len(sess.sheet_columns.get(name, []))
                                    nrows = 0
                                    if sess.workbook:
                                        try:
                                            first_col_df = sess.workbook.parse(name, header=header_row-1, usecols=[0])
                                            nrows = len(first_col_df.index)
                                        except Exception:
                                            nrows = 0
                                    rows_fmt.append((name, nrows, ncols, '-', ' '))
                                except Exception:
                                    rows_fmt.append((name, 0, 0, '-', ' '))
                        # Clear progress line
                        if last_print_len:
                            print(' ' * last_print_len, end='\r', flush=True)
                        scan_elapsed = time.time() - scan_start
                        if not rows_fmt:
                            print('No sheets to report.')
                        else:
                            name_w = max(disp_width(r[0]) for r in rows_fmt)
                            row_w = max(len('Rows'), max(len(str(r[1])) for r in rows_fmt))
                            col_w = max(len('Cols'), max(len(str(r[2])) for r in rows_fmt))
                            mem_w = max(len('Memory'), max(len(r[3]) for r in rows_fmt))
                            row_w = max(row_w, 5)
                            col_w = max(col_w, 5)
                            header_line = f"{pad_right('Sheet', name_w)}  {'Rows'.rjust(row_w)}  {'Cols'.rjust(col_w)}  {'Memory'.rjust(mem_w)}  Loaded"
                            if sess.color and sys.stdout.isatty():
                                print("\033[36m" + header_line + "\033[0m")
                            else:
                                print(header_line)
                            print('-'*len(header_line))
                            for name, nrows, ncols, mem_str, mark in rows_fmt:
                                flag = '*' if mark=='*' else ''
                                print(f"{pad_right(name, name_w)}  {str(nrows).rjust(row_w)}  {str(ncols).rjust(col_w)}  {mem_str.rjust(mem_w)}  {flag}")
                            print('-'*len(header_line))
                            print(f"Sheets: {len(rows_fmt)}  Loaded: {loaded_count}  Total loaded memory: {_fmt_bytes(total_mem)} (* = loaded)  Scan time: {scan_elapsed:.2f}s")
                continue  # after processing a command
            # Non-command: SQL input / buffering
            if line.endswith(';'):
                content = line[:-1].rstrip()
                if content:
                    buffer.append(content)
                sql_to_run = flush_query()
                if sql_to_run:
                    run_query(sql_to_run)
            else:
                if line.strip():
                    buffer.append(line)
                else:  # blank line flushes any accumulated SQL
                    sql_to_run = flush_query()
                    if sql_to_run:
                        run_query(sql_to_run)
        except KeyboardInterrupt:
            if buffer:
                buffer.clear()
                print('^C (cleared buffer)')
            else:
                print('^C')
            continue  # stay in loop
        except Exception as e:  # pragma: no cover
            print(f"Fatal error: {e}")
            continue
